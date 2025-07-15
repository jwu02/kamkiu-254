from datetime import datetime
import os
import random
import sys
import pandas as pd
from openpyxl import load_workbook

from errors import (
    ReportExistsError,
    CPKNotFoundError,
)

from utilities import (
    find_files_with_substrings,
    condense_row,
    get_mechanical_electrical_df_mask,
    get_metallographic_df_mask,
)

from constants import (
    MODEL_CODE_MAPPINGS,
    REPORT_OUTPUT_PATH,
    OQC_RETENTION_SAMPLE_CODES,
    PART_NAME,
    SCHEMA_CODE,
    CUSTOMER_PART_CODE,
)

class ShipmentBatch:
    def __init__(self, row: pd.Series):
        self.location = row['地区']
        self.customer = row['客户']
        self.shipment_date = row['发货日期']
        self.batch_quantity = row['发货数']

        self.model_code = row['型号']
        self.extrusion_batch_code = row['挤压批号']
        self.casting_furnace_code = row['炉号'] # 熔铸炉号
        self.ageing_batch_code = row['时效批号']

        self.die_code = row['模号']
        self.ageing_furnace_code = row['时效炉']
        self.extrusion_batch_qr_code_full = row['挤压批（二维码）']
        self.extrusion_batch_qr_code_half = row['挤压批次二维码']
        self.smelting_batch_code = row['熔铸批号']
        self.ageing_batch_qrcode = row['时效批次（二维码）']
        
        self.schema_code = row['图号']
        self.customer_part_code = row['客户料号']
        self.customer_batch_code = None

        self.project = row['项目']
        self.alloy_code = row['合金']
        self.recyle_ratio = row['回收比']

    def get_report_filename(self, total_batch_quantity: int) -> str:
        return f"{self.customer}MANCHESTER {self.model_code} {self.customer_part_code} {total_batch_quantity} ({self.location}) {self.casting_furnace_code} {self.extrusion_batch_code}"

    def calculate_batch_quantity_by_furnace_code(self, df_shipment_batch: pd.DataFrame) -> int:
        df_filtered = df_shipment_batch[
            (df_shipment_batch['地区'] == self.location) &
            (df_shipment_batch['客户'] == self.customer) &
            (df_shipment_batch['型号'] == self.model_code) &
            (df_shipment_batch['炉号'] == self.casting_furnace_code)
        ]

        return df_filtered['发货数'].astype(int).sum()

    def generate_report(
        self, 
        df_shipment_batch: pd.DataFrame,
        df_chemical_composition: pd.DataFrame,
        df_chemical_composition_limits: pd.DataFrame,
        df_functional_properties: pd.DataFrame,
        mid_plate_report_functional_requirements: pd.DataFrame,
        u_part_report_functional_requirements: pd.DataFrame,
    ) -> str:
        """
        报告模板生成函数
        填：型号、出货日期、图号、炉号、批量（出货数）、客户料号
        - CPK：查看对应的型号的CPK路径，再找对应挤压批号的CPK，复制数据过去模板
        """
        # print(f'{self.model_code} {self.casting_furnace_code} {self.ageing_batch_code}')

        total_batch_quantity = self.calculate_batch_quantity_by_furnace_code(df_shipment_batch)

        existing_report_files = find_files_with_substrings(REPORT_OUTPUT_PATH, [self.model_code, self.casting_furnace_code, self.location, self.customer])
        if existing_report_files:
            raise FileExistsError(f"报告已存在 {self.model_code} {self.casting_furnace_code}")

        if df_functional_properties is None:
            raise ValueError("未上传经过孤独 时效批号 和 熔铸炉号 搜索的性能数据")

        # Define template and output paths
        template_file = f"./报告模板/{self.model_code}.xlsx"  # Original template
        
        # Load the template workbook
        wb = load_workbook(template_file)
        ws = wb.active
        
        ws = self.report_basic_information(ws, total_batch_quantity)
        ws = self.report_cpk(ws)
        ws = self.report_functional_properties(
            ws, 
            df_functional_properties,
            mid_plate_report_functional_requirements, 
            u_part_report_functional_requirements
        )
        ws = self.report_chemical_composition(ws, df_chemical_composition, df_chemical_composition_limits)
        ws = self.report_weight(ws)

        output_name = self.get_report_filename(total_batch_quantity)
        output_file = os.path.join(REPORT_OUTPUT_PATH, f"{output_name}.xlsx")
        
        # Check if output directory exists, create if not
        os.makedirs(REPORT_OUTPUT_PATH, exist_ok=True)

        wb.save(output_file) # Save as new file
    
        return output_file
    
    def report_basic_information(self, ws, total_batch_quantity: int):
        # Fill in basic report information
        ws.cell(row=3, column=14, value=self.model_code) # 型号
        ws.cell(row=4, column=3, value=self.format_date(self.shipment_date)) # 发货日期
        ws.cell(row=4, column=7, value=SCHEMA_CODE[self.model_code]) # 图号
        ws.cell(row=4, column=11, value=self.casting_furnace_code) # 炉号
        ws.cell(row=4, column=14, value=total_batch_quantity) # 发货数
        ws.cell(row=4, column=19, value=CUSTOMER_PART_CODE[self.model_code]) # 客户料号
        ws.cell(row=5, column=3, value=PART_NAME[self.model_code]) # 品名
        ws.cell(row=8, column=1, value=self.calculate_sample_size(total_batch_quantity)) # 抽样数量
        
        return ws
    
    def report_cpk(self, ws):
        # Check if CPK file corresponding to model code exists
        cpk_path_str = MODEL_CODE_MAPPINGS[self.model_code]['cpk']['path']
        cpk_file_matches = find_files_with_substrings(cpk_path_str, [self.extrusion_batch_code])
        if not cpk_file_matches:
            raise FileNotFoundError(f"CPK不存在 {self.model_code} {self.casting_furnace_code}")
        
        cpk_path = os.path.join(cpk_path_str, cpk_file_matches[0])

        num_rows_to_extract = MODEL_CODE_MAPPINGS[self.model_code]['cpk']['num_rows']

        # Read data from existing CPK datasheet
        df_cpk_datasheet = pd.read_excel(
            cpk_path,
            engine='openpyxl',
            header=None,  # No headers (since we're reading raw cells)
            usecols="AH:AT",  # Columns from AH to AT
            skiprows=10,  # Skip first 10 rows (to start at row 11)
            nrows=num_rows_to_extract,  # Read x rows
        )
        
        # # Check if the DataFrame fits in the target range
        # if df.shape[0] > 64 or df.shape[1] > 11:
        #     show_error("Extracted data is too large for the target range")
        #     return
        
        # Write CPK data to report template
        for r_idx, row_data in enumerate(df_cpk_datasheet.values, start=12):  # Start at row 12
            for c_idx, value in enumerate(row_data, start=9):  # Start at column I (9)
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        return ws
    
    def report_functional_properties(
        self,
        ws,
        df_functional_properties: pd.DataFrame,
        mid_plate_report_functional_requirements: pd.DataFrame,
        u_part_report_functional_requirements: pd.DataFrame
    ):
        # Read functional requirements for both mid plate and u part parts, containing test type and test points
        if self.model_code == 'KAP-7461中板-A76-50':
            row_headings_dataframe = mid_plate_report_functional_requirements
        else:
            row_headings_dataframe = u_part_report_functional_requirements

        # Step 2: Create a row key
        row_headings_dataframe['row_key'] = row_headings_dataframe.apply(
            lambda row2: f"{row2['检测项目']}|{row2['项目详细']}|{row2['点位']}" if pd.notna(row2['点位'])
                        else f"{row2['检测项目']}|{row2['项目详细']}",
            axis=1
        )

        # Create separate parts each using different filter conditions
        df_row_headings_splitted = [
            {
                'df_row_heading': row_headings_dataframe[row_headings_dataframe['检测项目'].isin(['维氏硬度', '电导率', '室温拉伸'])],
                'additional_filter': (df_functional_properties['时效炉号'] == self.ageing_batch_code),
                'df_mask_function': get_mechanical_electrical_df_mask,
            },
            {
                'df_row_heading': row_headings_dataframe[row_headings_dataframe['检测项目'] == '铝合金金相显微组织'],
                'additional_filter': (df_functional_properties['铝棒炉号'] == self.casting_furnace_code),
                'df_mask_function': get_metallographic_df_mask,
            }
        ]

        # Step 3: Set up result DataFrame with full set of all sample codes from all groups
        all_columns = [code for group in OQC_RETENTION_SAMPLE_CODES for code in group]
        all_results = []

        for i in range(len(df_row_headings_splitted)):
            df_result = pd.DataFrame(index=df_row_headings_splitted[i]['df_row_heading']['row_key'], columns=all_columns)

            # Step 4: Fill in values incrementally, without overwriting existing ones
            for sample_code_group in OQC_RETENTION_SAMPLE_CODES:
                for _, row2 in df_row_headings_splitted[i]['df_row_heading'].iterrows():
                    test_group = row2['检测项目']
                    test_detail = row2['项目详细']
                    point = row2['点位']
                    row_key = row2['row_key']

                    for sample_code in sample_code_group:
                        # Skip if value already filled
                        if pd.notna(df_result.at[row_key, sample_code]):
                            continue

                        # Apply filter
                        condition = (
                            (df_functional_properties['检测项目'] == test_group) &
                            (df_functional_properties['oqc样号'] == sample_code) &
                            (df_functional_properties['型号'] == self.model_code) &
                            df_row_headings_splitted[i]['additional_filter']
                        )

                        alternative_condition = condition.copy()

                        if pd.notna(point):
                            condition &= (df_functional_properties['点位'] == point)
                        else:
                            # Stretch tests don't have test points
                            condition &= (df_functional_properties['点位'].isna())

                        matched = df_functional_properties[condition]

                        if not matched.empty:
                            value = matched.iloc[0].get(test_detail)
                            df_result.at[row_key, sample_code] = value
                        else:
                            # BRUTE FORCE APPROACH: Manually add alternative points
                            if test_group=='维氏硬度' and point=='C2':
                                print(f"{test_group} {point}")
                                alternative_condition &= (df_functional_properties['点位'] == 'S1')
                                matched = df_functional_properties[alternative_condition]
                                if not matched.empty:
                                    value = matched.iloc[0].get(test_detail)
                                    df_result.at[row_key, sample_code] = value
                            
                            if test_group=='铝合金金相显微组织' and point=='S7':
                                print(f"{test_group} {point}")
                                alternative_condition &= (df_functional_properties['点位'] == 'S10')
                                matched = df_functional_properties[alternative_condition]
                                if not matched.empty:
                                    value = matched.iloc[0].get(test_detail)
                                    df_result.at[row_key, sample_code] = value
                            
                            if test_group=='铝合金金相显微组织' and point=='S8':
                                print(f"{test_group} {point}")
                                alternative_condition &= (df_functional_properties['点位'] == 'S13')
                                matched = df_functional_properties[alternative_condition]
                                if not matched.empty:
                                    value = matched.iloc[0].get(test_detail)
                                    df_result.at[row_key, sample_code] = value
                            
                            if test_group=='铝合金金相显微组织' and point=='S9':
                                alternative_condition &= (df_functional_properties['点位'] == 'S16')
                                matched = df_functional_properties[alternative_condition]
                                if not matched.empty:
                                    value = matched.iloc[0].get(test_detail)
                                    df_result.at[row_key, sample_code] = value


            df_result = df_result.apply(condense_row, axis=1, result_type='expand')
            df_result = df_result.head(len(df_row_headings_splitted[i]['df_row_heading'])).iloc[:, :4]
            df_result = df_result.where(df_row_headings_splitted[i]['df_mask_function'](df_result), pd.NA)

            all_results.append(df_result)

        # Concatenate vertically (stack rows)
        df_extracted_functional_properties = pd.concat(all_results, axis=0)
        
        # Write functional properties to report template
        start_row = MODEL_CODE_MAPPINGS[self.model_code]['性能']['start_row']
        start_column = MODEL_CODE_MAPPINGS[self.model_code]['性能']['start_column']
        for r_idx, row_data in enumerate(df_extracted_functional_properties.values, start=start_row):
            for c_idx, value in enumerate(row_data, start=start_column):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        return ws
    
    def report_chemical_composition(
        self,
        ws,
        df_chemical_composition: pd.DataFrame,
        df_chemical_composition_limits: pd.DataFrame
    ):
        # Write chemical composition data to report template
        compositions = df_chemical_composition_limits['成分'].tolist()
        df_chemical_composition_filtered = df_chemical_composition[df_chemical_composition['炉号'] == self.casting_furnace_code]

        if df_chemical_composition_filtered.empty:
            raise ValueError(f"找不到对应炉号 {self.casting_furnace_code} 的化学成分数据")
        else:
            composition_data = df_chemical_composition_filtered.iloc[0].reindex(compositions)
            for r_idx, value in enumerate(composition_data, start=MODEL_CODE_MAPPINGS[self.model_code]['composition']['start_row']):
                ws.cell(row=r_idx, column=MODEL_CODE_MAPPINGS[self.model_code]['composition']['start_column'], value=round(float(value), 4))
        
        return ws
    
    def report_weight(self, ws):
        weights = self.generate_random_weights(self.model_code)
        # Weight cell starting values / coordinates
        weight_starting_row = MODEL_CODE_MAPPINGS[self.model_code]['重量']['starting_row']
        weight_starting_column = MODEL_CODE_MAPPINGS[self.model_code]['重量']['starting_column']
        for c_index, weight_val in enumerate(weights, start=weight_starting_column):
            ws.cell(row=weight_starting_row, column=c_index, value=weight_val)
        
        return ws
    
    def generate_random_weights(self, model_code: str) -> list[str]:
        lower_limit = MODEL_CODE_MAPPINGS[model_code]['重量']['lower_limit']
        upper_limit = MODEL_CODE_MAPPINGS[model_code]['重量']['upper_limit']

        weights = []

        for i in range(3):
            w = round(random.randint(lower_limit, upper_limit)/10, 3)
            weights.append(f'{w}g')

        return weights
    
    def format_date(self, s: str) -> str:
        # Replace hyphens with slashes to normalize
        normalized = s.replace("-", "/")
        
        try:
            # Parse the date
            dt = datetime.strptime(normalized, "%Y/%m/%d")
            
            # Get components as integers to remove leading zeros
            year = dt.year
            month = dt.month
            day = dt.day
            
            # Format without leading zeros
            return f"{year}/{month}/{day}"
        except ValueError:
            raise ValueError(f"Date '{s}' doesn't match expected formats (YYYY-MM-DD or YYYY/MM/DD)")
    
    def calculate_sample_size(self, batch_size: int) -> str:
        sample_size = 0
        if batch_size < 2: return "不够数量"
        elif batch_size <= 32: return "全检"
        elif batch_size <= 500: sample_size = 32
        elif batch_size <= 3200: sample_size = 125
        elif batch_size <= 10000: sample_size = 200
        elif batch_size <= 35000: sample_size = 315
        elif batch_size <= 150000: sample_size = 500
        elif batch_size <= 500000: sample_size = 800
        else: sample_size = 1250

        return f"{sample_size} pcs"