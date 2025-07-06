import os
import sys
import pandas as pd
from openpyxl import load_workbook
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QPushButton, QHBoxLayout, QVBoxLayout,
    QFileDialog, QTableWidgetItem, QLabel, QMessageBox
)

from MultiSelectionTable import MultiSelectionTable

from utilities import (
    transform_extrusion_batch_code,
    CheckStatus,
    SAMPLE_TYPES,
    load_cpk_tolerance_map,
    check_cpk_conformance,
    MODEL_CODE_MAPPINGS,
    find_files_with_substring,
    get_report_name,
    MODEL_CODE_ORDER,
    extract_location,
    extract_customer,
)

# https://pandas.pydata.org/pandas-docs/stable/user_guide/indexing.html#returning-a-view-versus-a-copy
pd.options.mode.copy_on_write = True

class KamKiu254(QMainWindow):
    DOCUMENT_NOT_UPLOADED = ""
    UPLOAD_CSV = "上传 CSV"

    def __init__(self):
        super().__init__()

        self.df_shipment_batch = None
        self.df_chemical_composition = None
        self.df_ageing_qrcode = None
        self.df_process_card_qrcode = None

        self.cpk_tolerance_map = load_cpk_tolerance_map()

        self.init_ui()
        

    def init_ui(self):
        """
        创造界面
        """
        self.setWindowTitle("254 发货报告主动生成")
        # self.resize(800, 600)

        # 上传 发货批次表
        self.shipment_batch_layout = QHBoxLayout()
        self.shipment_batch_layout.addWidget(QLabel("发货批次表："))
        self.shipment_path_label = QLabel(self.DOCUMENT_NOT_UPLOADED)
        self.shipment_batch_layout.addWidget(self.shipment_path_label)
        self.shipment_upload_button = QPushButton(self.UPLOAD_CSV)
        self.shipment_upload_button.clicked.connect(self.upload_shipment_batch_csv)
        self.shipment_batch_layout.addWidget(self.shipment_upload_button)
        self.shipment_batch_layout.addStretch()

        # 上传 化学成分
        self.composition_layout = QHBoxLayout()
        self.composition_layout.addWidget(QLabel("化学成分："))
        self.composition_path_label = QLabel(self.DOCUMENT_NOT_UPLOADED)
        self.composition_layout.addWidget(self.composition_path_label)
        self.composition_upload_button = QPushButton(self.UPLOAD_CSV)
        self.composition_upload_button.clicked.connect(self.upload_chemical_composition_csv)
        self.composition_layout.addWidget(self.composition_upload_button)
        self.composition_layout.addStretch()

        # 上传 型材时效二维码
        self.ageing_qrcode_layout = QHBoxLayout()
        self.ageing_qrcode_layout.addWidget(QLabel("型材时效二维码："))
        self.ageing_qrcode_path_label = QLabel(self.DOCUMENT_NOT_UPLOADED)
        self.ageing_qrcode_layout.addWidget(self.ageing_qrcode_path_label)
        self.ageing_qrcode_upload_button = QPushButton(self.UPLOAD_CSV)
        self.ageing_qrcode_upload_button.clicked.connect(self.upload_ageing_qrcode_csv)
        self.ageing_qrcode_layout.addWidget(self.ageing_qrcode_upload_button)
        self.ageing_qrcode_layout.addStretch()

        # 上传 流程卡二维码记录
        self.process_card_qrcode_layout = QHBoxLayout()
        self.process_card_qrcode_layout.addWidget(QLabel("流程卡二维码记录："))
        self.process_card_qrcode_path_label = QLabel(self.DOCUMENT_NOT_UPLOADED)
        self.process_card_qrcode_layout.addWidget(self.process_card_qrcode_path_label)
        self.process_card_qrcode_upload_button = QPushButton(self.UPLOAD_CSV)
        self.process_card_qrcode_upload_button.clicked.connect(self.upload_process_card_qrcode_csv)
        self.process_card_qrcode_layout.addWidget(self.process_card_qrcode_upload_button)
        self.process_card_qrcode_layout.addStretch()

        # 其他功能
        self.other_functionalities_layout = QHBoxLayout()
        # 检查CPk
        self.check_cpk_button = QPushButton("检查CPK")
        self.check_cpk_button.clicked.connect(self.check_cpk_path)
        self.other_functionalities_layout.addWidget(self.check_cpk_button)
        # 检查化学成分
        self.check_chemical_compositions_button = QPushButton("检查化学成分")
        self.check_chemical_compositions_button.clicked.connect(self.check_chemical_composition_conformance)
        self.other_functionalities_layout.addWidget(self.check_chemical_compositions_button)
        # 采取 挤压批此（二维码） & 熔铸批号
        self.fill_extrusion_batch_button = QPushButton("采取 挤压批此（二维码） & 熔铸批号")
        self.fill_extrusion_batch_button.clicked.connect(self.extract_data_from_ageing_qrcode)
        self.other_functionalities_layout.addWidget(self.fill_extrusion_batch_button)
        # 采取 时效批次（二维码）
        self.fill_ageing_batch_button = QPushButton("采取 时效批次（二维码）")
        self.fill_ageing_batch_button.clicked.connect(self.extract_ageing_batch_qrcode)
        self.other_functionalities_layout.addWidget(self.fill_ageing_batch_button)

        self.main_table = MultiSelectionTable()

        layout = QVBoxLayout()
        layout.addLayout(self.shipment_batch_layout)
        layout.addLayout(self.composition_layout)
        layout.addLayout(self.ageing_qrcode_layout)
        layout.addLayout(self.process_card_qrcode_layout)
        layout.addLayout(self.other_functionalities_layout)
        layout.addWidget(self.main_table)

        # Main widget and layout
        main_widget = QWidget()
        main_widget.setLayout(layout)
        self.setCentralWidget(main_widget)


    def upload_shipment_batch_csv(self):
        """
        上传 发货批次表
        """
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open CSV File", "", "CSV Files (*.csv)"
        )

        if file_path:
            self.shipment_path_label.setText(f"{file_path}")
            try:
                df = pd.read_csv(file_path)

                self.df_shipment_batch = df.reindex(columns=[
                    '地区',
                    '项目',
                    '发货数',
                    '发货日期',
                    '型号',
                    '挤压批号', # 第二列 挤压批号（有两列）
                    '挤压批（二维码）',
                    '炉号',
                    '熔铸批号',
                    '时效批号',
                    '时效批次（二维码）',
                    '客户'
                ])
                
                # Apply to DataFrame
                self.df_shipment_batch['时效批号（sfc）'] = self.df_shipment_batch['时效批号'].apply(lambda x: x+'*')

                # self.df_shipment_batch['客户/地区'] = self.df_shipment_batch['客户/地区'].apply(normalize_group_key)
                self.df_shipment_batch['地区'] = df['客户/地区'].apply(extract_location)
                self.df_shipment_batch['客户'] = df['客户/地区'].apply(extract_customer)
                self.df_shipment_batch['挤压批号'] = self.df_shipment_batch['挤压批号'].apply(transform_extrusion_batch_code)

                self.df_shipment_batch['型号'] = pd.Categorical(
                    self.df_shipment_batch['型号'], 
                    categories=MODEL_CODE_ORDER, 
                    ordered=True
                )
                self.df_shipment_batch.sort_values(by=['地区', '客户', '型号', '炉号', '发货数', '挤压批号', '时效批号'], inplace=True)
                self.df_shipment_batch.reset_index(drop=True, inplace=True)

                self.df_shipment_batch['CPK'] = CheckStatus.NOT_CHECKED.value
                self.df_shipment_batch['性能'] = CheckStatus.NOT_CHECKED.value
                self.df_shipment_batch['成分'] = CheckStatus.NOT_CHECKED.value

                self.display_dataframe(self.df_shipment_batch)
                self.display_report_generation_buttons()
                
            except Exception as e:
                self.shipment_path_label.setText(f"读取文件出错: {str(e)}")

    def display_dataframe(self, df):
        """
        显示 数据框架
        """
        self.main_table.setRowCount(len(df.index))
        self.main_table.setColumnCount(len(df.columns))
        self.main_table.setHorizontalHeaderLabels(df.columns.astype(str).tolist())

        for row in range(len(df.index)):
            for col in range(len(df.columns)):
                item = QTableWidgetItem(str(df.iat[row, col]))
                self.main_table.setItem(row, col, item)
        
        self.main_table.resizeColumnsToContents()
    
    def display_report_generation_buttons(self):
        num_table_cols = self.main_table.columnCount()
        self.main_table.insertColumn(num_table_cols)
        # Set the column header
        self.main_table.setHorizontalHeaderItem(num_table_cols, QTableWidgetItem("操作"))

        for index, row in self.df_shipment_batch.iterrows():
            # Create a button for the third column
            button = QPushButton('生成报告')
            button.clicked.connect(lambda _, i=index, r=row: self.generate_report(i, r))
            self.main_table.setCellWidget(index, num_table_cols, button)
    
    def generate_report(self, index, row):
        """
        报告模板生成函数
        填：型号、出货日期、图号、炉号、批量（出货数）、客户料号
        - CPK：查看对应的型号的CPK路径，再找对应挤压批号的CPK，复制数据过去模板
        - 
        """
        model_code = row['型号']
        extrusion_batch_code = row['挤压批号']

        cpk_path_str = MODEL_CODE_MAPPINGS[model_code]['cpk']['path']
        num_rows_to_extract = MODEL_CODE_MAPPINGS[model_code]['cpk']['num_rows']

        matching_files = find_files_with_substring(cpk_path_str, extrusion_batch_code)

        # print(f"===========INDEX {index}============")
        # print(row)

        file_count = len(matching_files)
        if file_count > 0:
            cpk_path = os.path.join(cpk_path_str, matching_files[0])

            # ================ get data from existing CPK files
            df = pd.read_excel(
                cpk_path,
                engine='openpyxl',
                header=None,  # No headers (since we're reading raw cells)
                usecols="AH:AT",  # Columns from AH to AT
                skiprows=10,  # Skip first 10 rows (to start at row 11)
                nrows=num_rows_to_extract,  # Read x rows
            )

            # Define template and output paths
            template_file = f"./report_templates/{model_code}.xlsx"  # Original template
            output_name = get_report_name(
                model_code,
                MODEL_CODE_MAPPINGS[model_code]['customer_part_code'],
                row['发货数'],
                'xx',
                row['炉号'],
                extrusion_batch_code
            )
            output_dir = r'.\output' # Where to save reports
            output_file = os.path.join(output_dir, f"{output_name}.xlsx")
            
            # Check if output directory exists, create if not
            os.makedirs(output_dir, exist_ok=True)
            
            # # Check if the DataFrame fits in the target range
            # if df.shape[0] > 64 or df.shape[1] > 11:
            #     show_error("Extracted data is too large for the target range")
            #     return
                
            # Load the template workbook
            wb = load_workbook(template_file)
            ws = wb.active  # or specify a specific sheet
            
            # Write DataFrame to target range
            for r_idx, row_data in enumerate(df.values, start=12):  # Start at row 12
                for c_idx, value in enumerate(row_data, start=9):  # Start at column I (9)
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Save as new file
            wb.save(output_file)
            show_error(f"Report successfully saved as {output_file}")
        else:
            show_error("对应的 CPK 不存在")
    
    
    def upload_chemical_composition_csv(self):
        """
        上传 化学成分
        """
        self.setDesiredElements()
        self.setDesiredSampleTypes()

        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open CSV File", "", "CSV Files (*.csv)"
        )

        if file_path:
            self.composition_path_label.setText(f"{file_path}")
            try:
                df = pd.read_csv(file_path)

                compositions = self.df_chemical_composition_limits['成分'].tolist()

                self.df_chemical_composition = df.reindex(columns=['炉号', '类型', *compositions])
                self.df_chemical_composition['Mn+Cr'] = 0
                self.df_chemical_composition = self.df_chemical_composition[self.df_chemical_composition['类型'].isin(SAMPLE_TYPES)]
                self.df_chemical_composition = self.df_chemical_composition.replace('-', pd.NA).dropna(how='any')

                self.df_chemical_composition.sort_values(by=['炉号', '类型'], ascending=[True, False], inplace=True)
                self.df_chemical_composition.reset_index(drop=True, inplace=True)
                self.df_chemical_composition['Mn+Cr'] = round(self.df_chemical_composition['Mn'].astype(float) + self.df_chemical_composition['Cr'].astype(float), 5)

                self.display_dataframe(self.df_chemical_composition)
                
            except Exception as e:
                self.composition_path_label.setText(f"读取文件出错: {str(e)}")
    
    def setDesiredElements(self):
        try:
            self.df_chemical_composition_limits = pd.read_csv("./data/成分_元素条件.csv")
            self.df_chemical_composition_limits = self.df_chemical_composition_limits.astype({ '上限': float, '下限': float })
        except Exception as e:
            show_error(f"读取文件出错: {str(e)}")
    
    def setDesiredSampleTypes(self):
        try:
            self.desired_sample_types = pd.read_csv("./data/成分_类型条件.csv")
        except Exception as e:
            show_error(f"读取文件出错: {str(e)}")

    def upload_ageing_qrcode_csv(self):
        """
        上传 型材时效二维码
        """
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open CSV File", "", "CSV Files (*.csv)"
        )

        if file_path:
            self.ageing_qrcode_path_label.setText(f"{file_path}")
            try:
                df = pd.read_csv(file_path)

                self.df_ageing_qrcode = df[[
                    '型号',
                    '生产挤压批',
                    '铝棒炉号',
                    '内部时效批',
                    '挤压批',
                    '熔铸批号',
                ]]

                self.df_ageing_qrcode.sort_values(by=['型号', '铝棒炉号', '生产挤压批'], inplace=True)
                self.df_ageing_qrcode.reset_index(drop=True, inplace=True)

                self.display_dataframe(self.df_ageing_qrcode)
                
            except Exception as e:
                self.ageing_qrcode_path_label.setText(f"读取文件出错: {str(e)}")
    
    def upload_process_card_qrcode_csv(self):
        """
        上传 流程卡二维码记录
        """
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open CSV File", "", "CSV Files (*.csv)"
        )

        if file_path:
            self.process_card_qrcode_path_label.setText(f"{file_path}")
            try:
                df = pd.read_csv(file_path)

                self.df_process_card_qrcode = df[[
                    '型号',
                    '挤压批号',
                    '炉号',
                    'sfc',
                    '二维码',
                ]]

                self.df_process_card_qrcode.rename(columns={'sfc': '时效批'}, inplace=True)
                self.df_process_card_qrcode['时效批'] = self.df_process_card_qrcode['时效批'].apply(lambda x: x[:8])

                self.display_dataframe(self.df_process_card_qrcode)
                
            except Exception as e:
                self.process_card_qrcode_path_label.setText(f"读取文件出错: {str(e)}")

    def check_single_record_cpk(self, index, row):
        error_path = []

        model_code = row['型号']
        extrusion_batch = str(row['挤压批号']).strip()
        path = MODEL_CODE_MAPPINGS[model_code]['cpk']['path']

        if not path or not os.path.isdir(path):
            self.df_shipment_batch.at[index, 'CPK'] = "🔴 错误"
            if path not in error_path:
                show_error(f"{model_code} 型号的路径找不到：${path}")
                error_path.append(path)
            return

        # Check if any file contains the extrusion batch string
        matching_files = find_files_with_substring(path, extrusion_batch)
        file_count = len(matching_files)

        if file_count == 0:
            self.df_shipment_batch.at[index, 'CPK'] = "🟠 不存在"
        else:
            # check CPK conformance
            if file_count > 1:
                self.df_shipment_batch.at[index, 'CPK'] = "🟠 多数CPK存在"
            else:
                # self.df_shipment_batch.at[index, 'CPK'] = "🟢 存在"

                file_path = os.path.join(path, matching_files[0])
                self.df_shipment_batch.at[index, 'CPK'] = check_cpk_conformance(file_path, self.cpk_tolerance_map[model_code])
            
    def check_cpk_path(self):
        if self.df_shipment_batch is None:
            show_error("请上传 发货批次表 数据")
            return

        for index, row in self.df_shipment_batch.iterrows():
            self.check_single_record_cpk(index, row)

        self.display_dataframe(self.df_shipment_batch)
        self.display_report_generation_buttons()
    
    def check_chemical_composition_conformance(self):
        """
        检查 化学成分
        """
        if self.df_shipment_batch is None:
            show_error("请上传 发货批次表 数据")
            return 
        if self.df_chemical_composition is None:
            show_error("请上传 化学成分 数据")
            return
        
        for index, row in self.df_shipment_batch.iterrows():
            furnace_code = row['炉号']

            # check if corresponding furnace
            df = self.df_chemical_composition[
                (self.df_chemical_composition['炉号'] == furnace_code)
            ]

            if len(df)>0:
                first_row = df.iloc[0]
                
                for index2, row2 in self.df_chemical_composition_limits.iterrows():
                    element = row2['成分']
                    upper_limit = row2['上限']
                    lower_limit = row2['下限']

                    value = float(first_row[element])
                    
                    if not (lower_limit <= value <= upper_limit):
                        self.df_shipment_batch.at[index, '成分'] = "🔴 不合格"
                        break
                    self.df_shipment_batch.at[index, '成分'] = "🟢 合格"
            else:
                self.df_shipment_batch.at[index, '成分'] = "🟠 找不到炉号"
        
        self.display_dataframe(self.df_shipment_batch)
        self.display_report_generation_buttons()


    def extract_data_from_ageing_qrcode(self):
        """
        填入 挤压批 & 熔铸批号 二维码
        """
        if self.df_shipment_batch is None:
            show_error("请上传 发货批次表 数据")
            return
        if self.df_ageing_qrcode is None:
            show_error("请上传 型材时效二维码 数据")
            return

        for index, row in self.df_shipment_batch.iterrows():
            model_code = row['型号']
            extrusion_batch_code = row['挤压批号']
            furnace_code = row['炉号']

            df = self.df_ageing_qrcode[
                (self.df_ageing_qrcode['型号'] == model_code)
                & (self.df_ageing_qrcode['生产挤压批'] == extrusion_batch_code)
                & (self.df_ageing_qrcode['铝棒炉号'] == furnace_code)
            ]

            self.df_shipment_batch.at[index, '挤压批（二维码）'] = df.iloc[0]['挤压批'] if len(df)>0 else "🟠 没记录"
            self.df_shipment_batch.at[index, '熔铸批号'] = df.iloc[0]['熔铸批号'][2:] if len(df)>0 else "🟠 没记录"
        
        self.display_dataframe(self.df_shipment_batch)
        self.display_report_generation_buttons()
    
    def extract_ageing_batch_qrcode(self):
        """
        从 流程卡二维码记录 采取 时效批次二维码
        """
        if self.df_shipment_batch is None:
            show_error("请上传 发货批次表 数据")
            return
        if self.df_process_card_qrcode is None:
            show_error("请上传 流程卡二维码记录 数据")
            return
            
        for index, row in self.df_shipment_batch.iterrows():
            model_code = row['型号']
            extrusion_batch_code = row['挤压批号']
            furnace_code = row['炉号']
            ageing_code = row['时效批号']

            df = self.df_process_card_qrcode[
                (self.df_process_card_qrcode['型号'] == model_code)
                & (self.df_process_card_qrcode['挤压批号'] == extrusion_batch_code)
                & (self.df_process_card_qrcode['炉号'] == furnace_code)
                & (self.df_process_card_qrcode['时效批'] == ageing_code)
            ]

            self.df_shipment_batch.at[index, '时效批次（二维码）'] = df.iloc[0]['二维码'][-4:] if len(df)>0 else "🟠 没记录"
        
        self.display_dataframe(self.df_shipment_batch)
        self.display_report_generation_buttons()


def show_error(msg: str):
    # Create and show a warning message box
    msg_box = QMessageBox()
    msg_box.setIcon(QMessageBox.Icon.Warning)
    msg_box.setText(msg)
    msg_box.setWindowTitle("注意")
    msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
    msg_box.exec()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = KamKiu254()
    window.showMaximized()
    sys.exit(app.exec())
