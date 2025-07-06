import os
import openpyxl
import warnings
from openpyxl.utils import get_column_letter

def convert_to_values(file_path):
    """
    Convert an Excel file to values only, overwriting the original file
    
    Args:
        file_path (str): Path to the input Excel file
    """
    try:
        # Ignore openpyxl warnings about workbook protection
        warnings.filterwarnings('ignore', category=UserWarning)
        
        # Load the workbook with formulas
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        # Process each worksheet
        for sheet in wb.worksheets:
            # Iterate through all cells in the worksheet
            for row in sheet.iter_rows():
                for cell in row:
                    # Handle both regular formulas and array formulas
                    if cell.data_type == 'f' or (hasattr(cell, 'array_formula') and cell.array_formula):
                        # Replace formula with its current value
                        cell.value = cell.value
        
        # Save the workbook back to the original file
        wb.save(file_path)
        print(f"Converted: {file_path} (overwritten)")
        return True
    
    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")
        return False

def process_directory(directory='.'):
    """
    Process all .xlsx files in a directory, overwriting originals
    
    Args:
        directory (str): Directory to search for Excel files
    """
    processed_count = 0
    error_count = 0
    
    print("\nExcel Files to Process:")
    print("----------------------")
    
    # First display all files that will be processed
    for filename in os.listdir(directory):
        if filename.endswith('.xlsx') and not filename.startswith('~$'):
            print(f"- {filename}")
    
    # Ask for confirmation
    confirm = input("\nWARNING: This will overwrite the original files. Continue? (y/n): ").strip().lower()
    if confirm != 'y':
        print("Operation cancelled.")
        return
    
    # Process files
    for filename in os.listdir(directory):
        if filename.endswith('.xlsx') and not filename.startswith('~$'):
            file_path = os.path.join(directory, filename)
            if convert_to_values(file_path):
                processed_count += 1
            else:
                error_count += 1
    
    print(f"\nProcessing complete! {processed_count} files converted, {error_count} errors")

if __name__ == '__main__':
    print("Excel to Values Converter (Overwrite Originals)")
    print("----------------------------------------------")
    directory = input("Enter directory path (or press Enter for current directory): ").strip() or '.'
    process_directory(directory)